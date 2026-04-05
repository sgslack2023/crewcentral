"""
Excel utilities for bulk upload/download functionality.
Handles Excel generation, parsing, and validation for all master data entities.
"""
import re
from io import BytesIO
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Customer, Branch, ServiceType, MoveType, RoomSize, SOURCE_CHOICES, STAGE_CHOICES
from users.models import CustomUser


# Entity configurations for each model type
ENTITY_CONFIGS = {
    'customer': {
        'model': Customer,
        'unique_field': 'email',
        'columns': [
            {'field': 'full_name', 'header': 'Full Name', 'required': True, 'help': 'Customer full name (required)'},
            {'field': 'email', 'header': 'Email', 'required': True, 'help': 'Email address (required, must be unique)'},
            {'field': 'phone', 'header': 'Phone', 'required': False, 'help': 'Phone number'},
            {'field': 'company', 'header': 'Company', 'required': False, 'help': 'Company name'},
            {'field': 'address', 'header': 'Address', 'required': False, 'help': 'Street address'},
            {'field': 'city', 'header': 'City', 'required': False, 'help': 'City'},
            {'field': 'state', 'header': 'State/Province', 'required': False, 'help': 'State or Province'},
            {'field': 'country', 'header': 'Country', 'required': False, 'help': 'Country'},
            {'field': 'postal_code', 'header': 'Postal Code', 'required': False, 'help': 'Postal/ZIP code'},
            {'field': 'origin_address', 'header': 'Origin Address', 'required': False, 'help': 'Move origin address'},
            {'field': 'destination_address', 'header': 'Destination Address', 'required': False, 'help': 'Move destination address'},
            {'field': 'move_date', 'header': 'Move Date', 'required': False, 'help': 'Move date (YYYY-MM-DD)'},
            {'field': 'source', 'header': 'Source', 'required': False, 'help': f"Lead source: {', '.join([c[0] for c in SOURCE_CHOICES])}"},
            {'field': 'stage', 'header': 'Stage', 'required': False, 'help': f"Customer stage: {', '.join([c[0] for c in STAGE_CHOICES])}"},
            {'field': 'service_type', 'header': 'Service Type', 'required': False, 'help': 'Service type name (must exist)', 'lookup': 'ServiceType'},
            {'field': 'move_size', 'header': 'Move Size', 'required': False, 'help': 'Room size name (must exist)', 'lookup': 'RoomSize'},
            {'field': 'branch', 'header': 'Branch', 'required': False, 'help': 'Branch name (must exist)', 'lookup': 'Branch'},
            {'field': 'assigned_to', 'header': 'Assigned To', 'required': False, 'help': 'Assigned user email (must exist)', 'lookup': 'User'},
            {'field': 'notes', 'header': 'Notes', 'required': False, 'help': 'Additional notes'},
        ],
        'choice_fields': {
            'source': dict(SOURCE_CHOICES),
            'stage': dict(STAGE_CHOICES),
        }
    },
    'branch': {
        'model': Branch,
        'unique_field': 'name',
        'columns': [
            {'field': 'name', 'header': 'Name', 'required': True, 'help': 'Branch name (required, must be unique)'},
            {'field': 'destination', 'header': 'Destination', 'required': False, 'help': 'Destination location'},
            {'field': 'dispatch_location', 'header': 'Dispatch Location', 'required': False, 'help': 'Dispatch location'},
            {'field': 'sales_tax_percentage', 'header': 'Sales Tax %', 'required': False, 'help': 'Sales tax percentage (e.g., 8.25)'},
            {'field': 'is_active', 'header': 'Is Active', 'required': False, 'help': 'Active status (TRUE/FALSE)'},
        ],
        'choice_fields': {}
    },
    'service_type': {
        'model': ServiceType,
        'unique_field': 'service_type',
        'columns': [
            {'field': 'service_type', 'header': 'Service Type', 'required': True, 'help': 'Service type name (required, must be unique)'},
            {'field': 'scaling_factor', 'header': 'Scaling Factor', 'required': False, 'help': 'Scaling factor (default 1.0)'},
            {'field': 'color', 'header': 'Color', 'required': False, 'help': 'Hex color code (e.g., #FF5733)'},
            {'field': 'estimate_content', 'header': 'Estimate Content', 'required': False, 'help': 'Additional estimate content'},
            {'field': 'enabled', 'header': 'Enabled', 'required': False, 'help': 'Enabled status (TRUE/FALSE)'},
        ],
        'choice_fields': {}
    },
    'move_type': {
        'model': MoveType,
        'unique_field': 'name',
        'columns': [
            {'field': 'name', 'header': 'Name', 'required': True, 'help': 'Move type name (required, must be unique)'},
            {'field': 'description', 'header': 'Description', 'required': False, 'help': 'Description'},
            {'field': 'cubic_feet', 'header': 'Cubic Feet', 'required': False, 'help': 'Cubic feet (default 0)'},
            {'field': 'weight', 'header': 'Weight', 'required': False, 'help': 'Weight in lbs (default 0)'},
            {'field': 'is_active', 'header': 'Is Active', 'required': False, 'help': 'Active status (TRUE/FALSE)'},
        ],
        'choice_fields': {}
    },
    'room_size': {
        'model': RoomSize,
        'unique_field': 'name',
        'columns': [
            {'field': 'name', 'header': 'Name', 'required': True, 'help': 'Room size name (required, must be unique)'},
            {'field': 'description', 'header': 'Description', 'required': False, 'help': 'Description'},
            {'field': 'cubic_feet', 'header': 'Cubic Feet', 'required': False, 'help': 'Cubic feet (default 0)'},
            {'field': 'weight', 'header': 'Weight', 'required': False, 'help': 'Weight in lbs (default 0)'},
            {'field': 'is_active', 'header': 'Is Active', 'required': False, 'help': 'Active status (TRUE/FALSE)'},
        ],
        'choice_fields': {}
    },
    'time_window': {
        'model': None,  # Will be set dynamically when needed
        'unique_field': 'name',
        'columns': [
            {'field': 'name', 'header': 'Name', 'required': True, 'help': 'Time window name (required, must be unique)'},
            {'field': 'start_time', 'header': 'Start Time', 'required': True, 'help': 'Start time (HH:MM format)'},
            {'field': 'end_time', 'header': 'End Time', 'required': True, 'help': 'End time (HH:MM format)'},
            {'field': 'description', 'header': 'Description', 'required': False, 'help': 'Description'},
            {'field': 'is_active', 'header': 'Is Active', 'required': False, 'help': 'Active status (TRUE/FALSE)'},
        ],
        'choice_fields': {}
    },
}


def generate_excel_template(entity_type, queryset, organization):
    """
    Generate an Excel file with headers and existing data from the queryset.
    
    Args:
        entity_type: String identifier for the entity (e.g., 'customer', 'branch')
        queryset: Django queryset of existing records
        organization: Organization instance for foreign key lookups
    
    Returns:
        BytesIO object containing the Excel file
    """
    config = ENTITY_CONFIGS.get(entity_type)
    if not config:
        raise ValueError(f"Unknown entity type: {entity_type}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = entity_type.replace('_', ' ').title()
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    help_font = Font(italic=True, color="666666", size=9)
    help_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    columns = config['columns']
    
    # Row 1: Headers
    for col_idx, col_config in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_config['header'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # Set column width based on header length
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_config['header']) + 5, 15)
    
    # Row 2: Help text
    for col_idx, col_config in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_config['help'])
        cell.font = help_font
        cell.fill = help_fill
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Data rows
    row_num = 3
    for record in queryset:
        for col_idx, col_config in enumerate(columns, start=1):
            field = col_config['field']
            value = getattr(record, field, None)
            
            # Handle foreign key fields - get the display value
            if col_config.get('lookup'):
                if value is not None:
                    if col_config['lookup'] == 'User':
                        value = value.email if hasattr(value, 'email') else str(value)
                    elif hasattr(value, 'name'):
                        value = value.name
                    elif hasattr(value, 'service_type'):
                        value = value.service_type
                    else:
                        value = str(value)
            
            # Handle date fields
            if isinstance(value, (datetime, date)):
                value = value.strftime('%Y-%m-%d')
            
            # Handle time fields
            if hasattr(value, 'strftime') and not isinstance(value, (datetime, date)):
                value = value.strftime('%H:%M')
            
            # Handle decimal fields
            if isinstance(value, Decimal):
                value = float(value)
            
            # Handle boolean fields
            if isinstance(value, bool):
                value = 'TRUE' if value else 'FALSE'
            
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = thin_border
        
        row_num += 1
    
    # Freeze the header rows
    ws.freeze_panes = 'A3'
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def parse_and_validate_excel(file, entity_type, organization, user=None, dry_run=True):
    """
    Parse and validate an Excel file for bulk import.
    
    Args:
        file: Uploaded file object
        entity_type: String identifier for the entity
        organization: Organization instance
        user: User instance (for created_by field when saving)
        dry_run: If True, only validate without saving to database
    
    Returns:
        Dictionary with validation results and optionally import results
    """
    config = ENTITY_CONFIGS.get(entity_type)
    if not config:
        return {'error': f"Unknown entity type: {entity_type}"}
    
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return {'error': f"Failed to read Excel file: {str(e)}"}
    
    columns = config['columns']
    unique_field = config['unique_field']
    model = config['model']
    choice_fields = config.get('choice_fields', {})
    
    # Handle dynamic model import for time_window
    if entity_type == 'time_window' and model is None:
        from transactiondata.models import TimeWindow
        model = TimeWindow
        config['model'] = model
    
    # Build header mapping (column index to field config)
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header_mapping = {}
    
    for col_idx, header_value in enumerate(header_row):
        if header_value:
            header_clean = str(header_value).strip().lower()
            for col_config in columns:
                if col_config['header'].lower() == header_clean:
                    header_mapping[col_idx] = col_config
                    break
    
    if not header_mapping:
        return {'error': "Could not find any matching column headers in the Excel file"}
    
    # Check required fields are present in headers
    required_fields = [c['field'] for c in columns if c['required']]
    mapped_fields = [header_mapping[idx]['field'] for idx in header_mapping]
    missing_required = [f for f in required_fields if f not in mapped_fields]
    
    if missing_required:
        return {'error': f"Missing required columns: {', '.join(missing_required)}"}
    
    # Track seen values for duplicate detection within file
    seen_unique_values = {}
    
    # Process data rows (skip header row 1 and help row 2)
    results = {
        'total': 0,
        'valid': 0,
        'warnings': 0,
        'errors': 0,
        'new_records': 0,
        'updates': 0,
        'rows': []
    }
    
    # For import mode
    created_count = 0
    updated_count = 0
    import_errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        
        results['total'] += 1
        row_data = {}
        row_issues = []
        row_status = 'valid'
        
        # Extract and validate each field
        for col_idx, col_config in header_mapping.items():
            field = col_config['field']
            raw_value = row[col_idx] if col_idx < len(row) else None
            
            # Clean the value
            if raw_value is not None:
                if isinstance(raw_value, str):
                    raw_value = raw_value.strip()
                    if raw_value == '':
                        raw_value = None
            
            row_data[field] = raw_value
            
            # Required field validation
            if col_config['required'] and raw_value is None:
                row_issues.append({
                    'field': field,
                    'message': f"{col_config['header']} is required",
                    'severity': 'error'
                })
                row_status = 'error'
        
        # Email format validation
        if 'email' in row_data and row_data.get('email'):
            email = row_data['email']
            email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_regex, email):
                row_issues.append({
                    'field': 'email',
                    'message': 'Invalid email format',
                    'severity': 'error'
                })
                row_status = 'error'
        
        # Choice field validation
        for choice_field, valid_choices in choice_fields.items():
            if choice_field in row_data and row_data.get(choice_field):
                value = str(row_data[choice_field]).lower()
                if value not in [k.lower() for k in valid_choices.keys()]:
                    row_issues.append({
                        'field': choice_field,
                        'message': f"Invalid value. Must be one of: {', '.join(valid_choices.keys())}",
                        'severity': 'error'
                    })
                    row_status = 'error'
                else:
                    # Normalize to correct case
                    for k in valid_choices.keys():
                        if k.lower() == value:
                            row_data[choice_field] = k
                            break
        
        # Date field parsing
        if 'move_date' in row_data and row_data.get('move_date'):
            date_value = row_data['move_date']
            parsed_date = None
            
            if isinstance(date_value, (datetime, date)):
                parsed_date = date_value if isinstance(date_value, date) else date_value.date()
            elif isinstance(date_value, str):
                date_formats = ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(date_value, fmt).date()
                        break
                    except ValueError:
                        continue
            
            if parsed_date:
                row_data['move_date'] = parsed_date
            else:
                row_issues.append({
                    'field': 'move_date',
                    'message': 'Invalid date format. Use YYYY-MM-DD',
                    'severity': 'error'
                })
                row_status = 'error'
        
        # Time field parsing (for time_window)
        for time_field in ['start_time', 'end_time']:
            if time_field in row_data and row_data.get(time_field):
                time_value = row_data[time_field]
                parsed_time = None
                
                if hasattr(time_value, 'hour'):  # Already a time object
                    parsed_time = time_value
                elif isinstance(time_value, str):
                    time_formats = ['%H:%M', '%H:%M:%S', '%I:%M %p', '%I:%M:%S %p']
                    for fmt in time_formats:
                        try:
                            parsed_time = datetime.strptime(time_value, fmt).time()
                            break
                        except ValueError:
                            continue
                
                if parsed_time:
                    row_data[time_field] = parsed_time
                else:
                    row_issues.append({
                        'field': time_field,
                        'message': 'Invalid time format. Use HH:MM',
                        'severity': 'error'
                    })
                    row_status = 'error'
        
        # Decimal field parsing
        for decimal_field in ['sales_tax_percentage', 'scaling_factor', 'cubic_feet', 'weight']:
            if decimal_field in row_data and row_data.get(decimal_field) is not None:
                try:
                    row_data[decimal_field] = Decimal(str(row_data[decimal_field]))
                except (InvalidOperation, ValueError):
                    row_issues.append({
                        'field': decimal_field,
                        'message': 'Invalid number format',
                        'severity': 'error'
                    })
                    row_status = 'error'
        
        # Boolean field parsing
        for bool_field in ['is_active', 'enabled']:
            if bool_field in row_data and row_data.get(bool_field) is not None:
                value = row_data[bool_field]
                if isinstance(value, bool):
                    pass  # Already boolean
                elif isinstance(value, str):
                    row_data[bool_field] = value.lower() in ['true', 'yes', '1', 'y']
                elif isinstance(value, (int, float)):
                    row_data[bool_field] = bool(value)
        
        # Duplicate detection within file
        unique_value = row_data.get(unique_field)
        if unique_value:
            unique_key = str(unique_value).lower()
            if unique_key in seen_unique_values:
                row_issues.append({
                    'field': unique_field,
                    'message': f"Duplicate {unique_field} found in row {seen_unique_values[unique_key]}",
                    'severity': 'error'
                })
                row_status = 'error'
            else:
                seen_unique_values[unique_key] = row_idx
        
        # Check if record exists (NEW vs UPDATE)
        action = 'create'
        existing_record = None
        if unique_value and row_status != 'error':
            lookup_kwargs = {unique_field: unique_value, 'organization': organization}
            if entity_type == 'customer':
                # Customer email is globally unique, not org-scoped
                lookup_kwargs = {'email': unique_value}
            
            try:
                existing_record = model.objects.filter(**lookup_kwargs).first()
                if existing_record:
                    action = 'update'
            except Exception:
                pass
        
        # Foreign key lookups
        lookup_errors = _validate_foreign_keys(row_data, columns, organization)
        for err in lookup_errors:
            row_issues.append(err)
            if err['severity'] == 'error':
                row_status = 'error'
        
        # Update counts
        if row_status == 'error':
            results['errors'] += 1
        elif row_status == 'warning':
            results['warnings'] += 1
            results['valid'] += 1
        else:
            results['valid'] += 1
        
        if action == 'create':
            results['new_records'] += 1
        else:
            results['updates'] += 1
        
        # Add to rows list
        results['rows'].append({
            'row_num': row_idx,
            'data': {k: str(v) if v is not None else '' for k, v in row_data.items()},
            'status': row_status,
            'action': action,
            'issues': row_issues
        })
        
        # If not dry run and no errors, save the record
        if not dry_run and row_status != 'error':
            try:
                save_data = _prepare_save_data(row_data, columns, organization, user)
                
                if existing_record:
                    # Update existing record
                    for key, value in save_data.items():
                        setattr(existing_record, key, value)
                    existing_record.save()
                    updated_count += 1
                else:
                    # Create new record
                    save_data['organization'] = organization
                    if user and hasattr(model, 'created_by'):
                        save_data['created_by'] = user
                    model.objects.create(**save_data)
                    created_count += 1
                    
            except Exception as e:
                import_errors.append({
                    'row_num': row_idx,
                    'message': str(e)
                })
    
    if not dry_run:
        return {
            'created': created_count,
            'updated': updated_count,
            'skipped': results['errors'],
            'errors': import_errors
        }
    
    return results


def _validate_foreign_keys(row_data, columns, organization):
    """
    Validate foreign key fields by looking up values in the database.
    
    Returns:
        List of error dictionaries
    """
    errors = []
    
    for col_config in columns:
        if not col_config.get('lookup'):
            continue
        
        field = col_config['field']
        value = row_data.get(field)
        
        if not value:
            continue
        
        lookup_type = col_config['lookup']
        found = None
        
        try:
            if lookup_type == 'ServiceType':
                found = ServiceType.objects.filter(
                    organization=organization,
                    service_type__iexact=value
                ).first()
                
            elif lookup_type == 'RoomSize':
                found = RoomSize.objects.filter(
                    organization=organization,
                    name__iexact=value
                ).first()
                
            elif lookup_type == 'Branch':
                found = Branch.objects.filter(
                    organization=organization,
                    name__iexact=value
                ).first()
                
            elif lookup_type == 'User':
                found = CustomUser.objects.filter(
                    organization=organization,
                    email__iexact=value
                ).first()
            
            if not found:
                errors.append({
                    'field': field,
                    'message': f"{lookup_type} '{value}' not found",
                    'severity': 'error'
                })
            else:
                # Store the found object for later use
                row_data[f'_{field}_obj'] = found
                
        except Exception as e:
            errors.append({
                'field': field,
                'message': f"Error looking up {lookup_type}: {str(e)}",
                'severity': 'error'
            })
    
    return errors


def _prepare_save_data(row_data, columns, organization, user):
    """
    Prepare data dictionary for saving to database.
    Converts lookup field values to actual foreign key objects.
    """
    save_data = {}
    
    for col_config in columns:
        field = col_config['field']
        
        if col_config.get('lookup'):
            # Use the looked-up object
            obj_key = f'_{field}_obj'
            if obj_key in row_data:
                save_data[field] = row_data[obj_key]
        elif field in row_data and row_data[field] is not None:
            save_data[field] = row_data[field]
    
    return save_data
